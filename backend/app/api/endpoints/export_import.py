from fastapi import APIRouter, Depends, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from app.db.database import get_db
from app.models.organization import Organization
from app.models.log import OrganizationLog

router = APIRouter()

def get_full_path(org_id: int, db: Session, cache: dict) -> str:
    if org_id in cache:
        return cache[org_id]
    org = db.query(Organization).filter(Organization.id == org_id).first()
    if not org:
        return ''
    if org.parent_id is None:
        cache[org_id] = org.name
    else:
        parent_path = get_full_path(org.parent_id, db, cache)
        cache[org_id] = f"{parent_path} / {org.name}"
    return cache[org_id]

@router.get("/export")
def export_excel(db: Session = Depends(get_db)):
    orgs = db.query(Organization).filter(Organization.deleted_at == None).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "組織電話簿"

    headers = ["ID", "完整路徑", "單位名稱", "上層單位ID", "自動電話", "警用電話", "鐵路電話", "傳真電話", "地址", "備註"]
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    path_cache = {}
    for row, org in enumerate(orgs, 2):
        full_path = get_full_path(org.id, db, path_cache)
        ws.cell(row=row, column=1, value=org.id)
        ws.cell(row=row, column=2, value=full_path)
        ws.cell(row=row, column=3, value=org.name)
        ws.cell(row=row, column=4, value=org.parent_id)
        ws.cell(row=row, column=5, value=org.phone_auto)
        ws.cell(row=row, column=6, value=org.phone_police)
        ws.cell(row=row, column=7, value=org.phone_railway)
        ws.cell(row=row, column=8, value=org.phone_fax)
        ws.cell(row=row, column=9, value=org.address)
        ws.cell(row=row, column=10, value=org.note)

    col_widths = [8, 40, 20, 12, 15, 15, 15, 15, 30, 20]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=org_directory.xlsx"}
    )

@router.post("/import")
def import_excel(
    file: UploadFile = File(...),
    changed_by: str = Form(...),
    db: Session = Depends(get_db)
):
    contents = file.file.read()
    wb = openpyxl.load_workbook(BytesIO(contents))
    ws = wb.active

    results = {"created": 0, "updated": 0, "errors": []}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[2]:
            continue

        name = str(row[2]).strip()
        parent_id = row[3] if row[3] else None
        phone_auto = str(row[4]).strip() if row[4] else None
        phone_police = str(row[5]).strip() if row[5] else None
        phone_railway = str(row[6]).strip() if row[6] else None
        phone_fax = str(row[7]).strip() if row[7] else None
        address = str(row[8]).strip() if row[8] else None
        note = str(row[9]).strip() if row[9] else None

        existing = db.query(Organization).filter(
            Organization.name == name,
            Organization.parent_id == parent_id,
            Organization.deleted_at == None
        ).first()

        if existing:
            fields = {
                "phone_auto": phone_auto,
                "phone_police": phone_police,
                "phone_railway": phone_railway,
                "phone_fax": phone_fax,
                "address": address,
                "note": note,
            }
            for field, new_val in fields.items():
                old_val = getattr(existing, field)
                if old_val != new_val:
                    db.add(OrganizationLog(
                        organization_id=existing.id,
                        action="update",
                        field_changed=field,
                        old_value=str(old_val) if old_val else None,
                        new_value=str(new_val) if new_val else None,
                        changed_by=changed_by,
                    ))
                    setattr(existing, field, new_val)
            results["updated"] += 1
        else:
            org = Organization(
                name=name,
                parent_id=parent_id,
                phone_auto=phone_auto,
                phone_police=phone_police,
                phone_railway=phone_railway,
                phone_fax=phone_fax,
                address=address,
                note=note,
            )
            db.add(org)
            db.flush()
            db.add(OrganizationLog(
                organization_id=org.id,
                action="create",
                changed_by=changed_by,
            ))
            results["created"] += 1

    db.commit()
    return {"success": True, "results": results}
